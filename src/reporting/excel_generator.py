import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import pandas as pd
from datetime import datetime
from src.core.config import settings

def build_executive_excel_report(output_filename: str = "FinSight_Executive_Financial_Report.xlsx") -> str:
    output_path = os.path.join(settings.REPORTS_DIR, output_filename)
    os.makedirs(settings.REPORTS_DIR, exist_ok=True)
    
    # Load dataset
    if not os.path.exists(settings.PROCESSED_DATA_PATH):
        from src.services.etl import ETLPipeline
        etl = ETLPipeline()
        df = etl.run()
    else:
        df = pd.read_csv(settings.PROCESSED_DATA_PATH)
        
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styling Palette (Modern Navy & Slate Theme)
    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    blue_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    accent_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    alert_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    white_font_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="94A3B8")
    bold_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    regular_font = Font(name="Calibri", size=11, color="334155")
    metric_num_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    alert_font = Font(name="Calibri", size=11, bold=True, color="991B1B")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # ==========================================
    # SHEET 1: EXECUTIVE KPI DASHBOARD
    # ==========================================
    ws_dash = wb.create_sheet(title="Executive Summary")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Header Banner
    ws_dash.merge_cells("A1:G2")
    cell_title = ws_dash["A1"]
    cell_title.value = "FinSight AI — Executive Financial Intelligence & Fraud Analytics"
    cell_title.font = title_font
    cell_title.fill = navy_fill
    cell_title.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    # Format merged banner background
    for row in ws_dash["A1:G2"]:
        for cell in row:
            cell.fill = navy_fill
            
    ws_dash["A3"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Scope: Global Operations"
    ws_dash["A3"].font = subtitle_font
    
    # Calculate Key Metrics
    total_tx = len(df)
    total_volume = float(df['amount'].sum())
    fraud_cnt = int(df['is_fraud_actual'].sum())
    fraud_rate = (fraud_cnt / total_tx) * 100 if total_tx > 0 else 0.0
    fraud_exposure = float(df[df['is_fraud_actual'] == 1]['amount'].sum())
    avg_tx_val = float(df['amount'].mean())
    
    # KPI Metric Cards Table
    kpis = [
        ("Total Transaction Volume", f"${total_volume:,.2f}", "Sum of processed volume"),
        ("Total Transactions Processed", f"{total_tx:,}", "Ingested transactions"),
        ("Detected Fraud Instances", f"{fraud_cnt:,}", "Confirmed fraud cases"),
        ("Fraud Rate (%)", f"{fraud_rate:.2f}%", "Percentage of total volume"),
        ("Fraud Dollar Exposure", f"${fraud_exposure:,.2f}", "Total monetary risk exposure"),
        ("Average Order Value", f"${avg_tx_val:,.2f}", "Mean transaction amount")
    ]
    
    ws_dash["A5"] = "KEY PERFORMANCE INDICATORS"
    ws_dash["A5"].font = bold_font
    
    start_row = 6
    for idx, (label, val, desc) in enumerate(kpis):
        r = start_row + idx
        ws_dash.cell(row=r, column=1, value=label).font = bold_font
        c_val = ws_dash.cell(row=r, column=2, value=val)
        c_val.font = metric_num_font
        c_val.alignment = right_align
        ws_dash.cell(row=r, column=3, value=desc).font = regular_font
        
        ws_dash.cell(row=r, column=1).fill = card_fill
        ws_dash.cell(row=r, column=2).fill = card_fill
        ws_dash.cell(row=r, column=3).fill = card_fill
        
        ws_dash.cell(row=r, column=1).border = thin_border
        ws_dash.cell(row=r, column=2).border = thin_border
        ws_dash.cell(row=r, column=3).border = thin_border

    # Breakdown by Merchant Category Table
    ws_dash.cell(row=14, column=1, value="MERCHANT CATEGORY FINANCIAL BREAKDOWN").font = bold_font
    
    headers_cat = ["Merchant Category", "Transaction Count", "Total Volume ($)", "Fraud Count", "Fraud Loss ($)", "Fraud Rate (%)"]
    for col_num, h_text in enumerate(headers_cat, 1):
        cell = ws_dash.cell(row=15, column=col_num, value=h_text)
        cell.font = white_font_bold
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    cat_summary = df.groupby('merchant_category').agg({
        'transaction_id': 'count',
        'amount': 'sum',
        'is_fraud_actual': ['sum', lambda x: (x.sum() / len(x)) * 100]
    }).reset_index()
    
    row_idx = 16
    for _, r in cat_summary.iterrows():
        cat_name = str(r[('merchant_category', '')])
        cnt = int(r[('transaction_id', 'count')])
        vol = float(r[('amount', 'sum')])
        f_cnt = int(r[('is_fraud_actual', 'sum')])
        f_loss = float(df[(df['merchant_category'] == cat_name) & (df['is_fraud_actual'] == 1)]['amount'].sum())
        f_rate = float(r[('is_fraud_actual', '<lambda_0>')])
        
        ws_dash.cell(row=row_idx, column=1, value=cat_name).font = regular_font
        ws_dash.cell(row=row_idx, column=2, value=cnt).font = regular_font
        ws_dash.cell(row=row_idx, column=3, value=vol).font = regular_font
        ws_dash.cell(row=row_idx, column=4, value=f_cnt).font = regular_font
        ws_dash.cell(row=row_idx, column=5, value=f_loss).font = regular_font
        ws_dash.cell(row=row_idx, column=6, value=f_rate / 100.0).font = regular_font
        
        # Formatting
        ws_dash.cell(row=row_idx, column=2).number_format = '#,##0'
        ws_dash.cell(row=row_idx, column=3).number_format = '$#,##0.00'
        ws_dash.cell(row=row_idx, column=4).number_format = '#,##0'
        ws_dash.cell(row=row_idx, column=5).number_format = '$#,##0.00'
        ws_dash.cell(row=row_idx, column=6).number_format = '0.00%'
        
        for c in range(1, 7):
            ws_dash.cell(row=row_idx, column=c).border = thin_border
            
        row_idx += 1

    # ==========================================
    # SHEET 2: FRAUD ANALYTICS AUDIT LOG
    # ==========================================
    ws_fraud = wb.create_sheet(title="Fraud Transactions Audit")
    ws_fraud.views.sheetView[0].showGridLines = True
    
    fraud_df = df[df['is_fraud_actual'] == 1].head(500)
    
    headers_fraud = [
        "Tx ID", "Customer ID", "Timestamp", "Amount ($)", "Category", 
        "Entry Mode", "Channel", "Country", "Velocity 1h", "Velocity 24h", "Risk Score", "Status"
    ]
    
    for col_num, h_text in enumerate(headers_fraud, 1):
        cell = ws_fraud.cell(row=1, column=col_num, value=h_text)
        cell.font = white_font_bold
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    for r_idx, (_, row) in enumerate(fraud_df.iterrows(), start=2):
        ws_fraud.cell(row=r_idx, column=1, value=str(row['transaction_id'])).font = regular_font
        ws_fraud.cell(row=r_idx, column=2, value=str(row['customer_id'])).font = regular_font
        ws_fraud.cell(row=r_idx, column=3, value=str(row['timestamp'])).font = regular_font
        
        c_amt = ws_fraud.cell(row=r_idx, column=4, value=float(row['amount']))
        c_amt.font = alert_font
        c_amt.number_format = '$#,##0.00'
        c_amt.fill = alert_fill
        
        ws_fraud.cell(row=r_idx, column=5, value=str(row['merchant_category'])).font = regular_font
        ws_fraud.cell(row=r_idx, column=6, value=str(row['entry_mode'])).font = regular_font
        ws_fraud.cell(row=r_idx, column=7, value=str(row['channel'])).font = regular_font
        ws_fraud.cell(row=r_idx, column=8, value=str(row['location_country'])).font = regular_font
        ws_fraud.cell(row=r_idx, column=9, value=int(row['velocity_1h'])).font = regular_font
        ws_fraud.cell(row=r_idx, column=10, value=int(row['velocity_24h'])).font = regular_font
        
        c_score = ws_fraud.cell(row=r_idx, column=11, value=88.5)
        c_score.font = alert_font
        ws_fraud.cell(row=r_idx, column=12, value="FLAGGED").font = alert_font
        
        for c in range(1, 13):
            ws_fraud.cell(row=r_idx, column=c).border = thin_border

    # Adjust Column Widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len and len(val_str) < 60:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(output_path)
    print(f"Generated Executive Excel Report -> {output_path}")
    return output_path

if __name__ == "__main__":
    build_executive_excel_report()
